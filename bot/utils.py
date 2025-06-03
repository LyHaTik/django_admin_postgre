import os
from datetime import datetime

from aiogram.enums import ChatMemberStatus
from openpyxl import Workbook, load_workbook

from config import ORDERS_FILE, CHANNEL_ID, bot


# Проверка подписки
async def check_subscribed(user_id: int) -> bool:
    try:
        member = await bot.get_chat_member(chat_id=CHANNEL_ID, user_id=user_id)
        return member.status != ChatMemberStatus.LEFT
    except Exception:
        return False


# Парсинг вызовов
async def parse_data(data: str) -> tuple:
    parts = data.split(":")
    return int(parts[1]), parts[2] if len(parts) > 2 else None


# Генератор номера заказа
async def create_order_id() -> int:
    return int(datetime.now().strftime("%Y%m%d%H%M%S"))


# Сохранение заказа в Excel
async def save_to_excel(order_data: list[dict]):
    headers = ["Order ID", "Дата", "User ID", "Товар", "Кол-во", "Сумма", "Валюта", "Описание"]

    if os.path.exists(ORDERS_FILE):
        wb = load_workbook(ORDERS_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    for order in order_data:
        row = [order.get(h, "") for h in headers]
        ws.append(row)
    
    wb.save(ORDERS_FILE)
