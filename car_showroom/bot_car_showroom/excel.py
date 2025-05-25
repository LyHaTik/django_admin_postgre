import os
from openpyxl import Workbook, load_workbook

from .config import ORDERS_FILE


async def save(order_data: list):
    if os.path.exists(ORDERS_FILE):
        wb = load_workbook(ORDERS_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Order ID", "Дата", "User ID", "Товар", "Кол-во", "Сумма (₽)", "Валюта", "Описание"])

    for order in order_data:
        ws.append(order)
    wb.save(ORDERS_FILE)
